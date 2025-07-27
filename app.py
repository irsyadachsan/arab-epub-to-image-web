# app.py
# Modul utama aplikasi Flask yang mengorkestrasi seluruh alur konversi ePub ke gambar dan pemrosesan AI.

from flask import Flask, render_template, request, send_from_directory, jsonify
import os # Untuk operasi sistem file seperti membuat direktori, menghapus file
import logging # Untuk mencatat informasi, peringatan, dan error
import shutil # Untuk operasi file tingkat tinggi, seperti menghapus direktori (shutil.rmtree)
import random # Untuk memilih gambar fallback secara acak
import re # Untuk operasi regex, digunakan dalam membersihkan prompt
import time # Untuk mengukur waktu proses
import openpyxl # Untuk membaca dan menulis file Excel (.xlsx)
from openpyxl import Workbook, load_workbook # Import spesifik dari openpyxl

# Import modul-modul inti proyek yang telah dikembangkan
import epub_processor # Modul untuk ekstraksi konten ePub dan chunking teks
import image_renderer # Modul untuk rendering gambar (halaman ePub dan gambar hasil LLM)
import llm_integrator # Modul untuk berinteraksi dengan Google Gemini API dan Hugging Face API

from bs4 import BeautifulSoup # Digunakan untuk membersihkan teks HTML dari ePub sebelum dikirim ke LLM

from PIL import Image # Digunakan oleh Pillow untuk membuat gambar default jika diperlukan

# Import untuk ROUGE Score
from rouge_score import rouge_scorer # <--- PASTIKAN INI ADA DI SINI

# Konfigurasi dasar logging untuk aplikasi
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Inisialisasi aplikasi Flask
app = Flask(__name__)

# --- Konfigurasi Folder Aplikasi ---
# Folder untuk menyimpan file ePub yang diunggah oleh pengguna
UPLOAD_FOLDER = 'uploads'
# Folder untuk menyimpan semua gambar yang dihasilkan (baik halaman ePub maupun gambar hasil AI)
GENERATED_IMAGES_FOLDER = 'generated_images'
# Nama file untuk log kinerja dalam format Excel
PERFORMANCE_LOG_FILE = 'performance_log.xlsx' 

# Mengatur konfigurasi Flask untuk folder-folder yang digunakan
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['GENERATED_IMAGES_FOLDER'] = GENERATED_IMAGES_FOLDER

# Memastikan folder-folder yang dibutuhkan ada. Jika belum ada, akan dibuat secara otomatis.
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(GENERATED_IMAGES_FOLDER, exist_ok=True)
# Sub-direktori untuk menyimpan aset ePub yang diekstrak sementara
os.makedirs(os.path.join(UPLOAD_FOLDER, 'epub_extracts'), exist_ok=True)
# Folder untuk menyimpan gambar latar belakang fallback yang sudah didesain
os.makedirs(os.path.join(app.root_path, 'static', 'images', 'fallback_ai_bgs'), exist_ok=True)
# Folder untuk menyimpan font kustom yang digunakan oleh Pillow
os.makedirs(os.path.join(app.root_path, 'fonts'), exist_ok=True)


# Inisialisasi konfigurasi Google Gemini API saat aplikasi Flask dimulai.
# Kunci API (GOOGLE_API_KEY) harus diatur sebagai variabel lingkungan sebelum menjalankan aplikasi.
try:
    llm_integrator.configure_gemini()
except ValueError as e:
    # Jika kunci API tidak ditemukan, catat error. Fitur LLM tidak akan berfungsi.
    logging.error(f"Gagal mengkonfigurasi Gemini API saat startup: {e}. Pastikan GOOGLE_API_KEY diatur di variabel lingkungan Anda.")


# --- Fungsi Bantu (Helper Functions) ---

def extract_background_color_from_prompt(prompt):
    """
    Mengekstrak permintaan warna latar belakang dari prompt pengguna.
    
    Args:
        prompt (str): Prompt asli yang diberikan oleh pengguna.
        
    Returns:
        tuple: (modified_prompt, requested_color_rgb)
               - modified_prompt (str): Prompt setelah permintaan warna dihapus.
               - requested_color_rgb (tuple): Nilai RGB warna (tuple 3 int) jika ditemukan, None jika tidak.
    """
    # Peta warna dasar yang dikenali dari prompt ke nilai RGB
    color_map = {
        "merah": (255, 0, 0), "biru": (0, 0, 255), "hijau": (0, 128, 0),
        "kuning": (255, 255, 0), "hitam": (0, 0, 0), "putih": (255, 255, 255),
        "oranye": (255, 165, 0), "ungu": (128, 0, 128), "abu-abu": (128, 128, 128),
        "coklat": (165, 42, 42),
    }
    
    prompt_lower = prompt.lower() # Konversi prompt ke huruf kecil untuk pencarian
    found_color = None # Inisialisasi warna yang ditemukan

    # Iterasi melalui peta warna untuk mencari permintaan warna dalam prompt
    for color_name, rgb_value in color_map.items():
        if f"background berwarna {color_name}" in prompt_lower or \
           f"latar belakang {color_name}" in prompt_lower or \
           f"background {color_name}" in prompt_lower:
            found_color = rgb_value
            logging.info(f"Permintaan warna latar belakang terdeteksi: {color_name} -> {rgb_value}")
            break # Hentikan pencarian setelah warna pertama ditemukan
    
    # Jika warna ditemukan, hapus frasa permintaan warna dari prompt asli
    # Ini penting agar LLM tidak mencoba menginterpretasikan permintaan warna sebagai instruksi teks
    if found_color:
        for color_name in color_map.keys():
            # Menggunakan regex untuk menghapus frasa permintaan warna (case-insensitive)
            prompt = re.sub(r"background berwarna " + re.escape(color_name), "", prompt, flags=re.IGNORECASE).strip()
            prompt = re.sub(r"latar belakang " + re.escape(color_name), "", prompt, flags=re.IGNORECASE).strip()
            prompt = re.sub(r"background " + re.escape(color_name), "", prompt, flags=re.IGNORECASE).strip()
        prompt = re.sub(r'\s+', ' ', prompt).strip() # Hapus spasi ganda yang mungkin muncul setelah penghapusan
    
    return prompt, found_color

def hitung_rouge_score(reference_text, generated_text):
    """
    Menghitung ROUGE-1 F1 Score antara teks referensi dan teks yang dihasilkan.
    
    Catatan: Dalam implementasi saat ini, 'reference_text' adalah prompt LLM.
    Untuk evaluasi ROUGE yang akurat, 'reference_text' seharusnya adalah ringkasan/terjemahan
    yang dibuat oleh manusia (human-written reference).
    """
    if not reference_text or not generated_text:
        return 0.0 # Mengembalikan 0 jika salah satu teks kosong
    scorer = rouge_scorer.RougeScorer(['rouge1'], use_stemmer=True)
    score = scorer.score(reference_text, generated_text)
    rouge1_f1 = score['rouge1'].fmeasure
    return round(rouge1_f1, 4) # Bulatkan ke 4 angka di belakang koma

def log_performance_data(timestamp, epub_filename, llm_prompt, llm_response_text, rouge_score, total_duration, num_epub_pages, num_chunks, status_message):
    """
    Mencatat data kinerja setiap proses konversi ke file Excel.
    
    Args:
        timestamp (str): Waktu proses dicatat.
        epub_filename (str): Nama file ePub yang diproses.
        llm_prompt (str): Prompt LLM asli yang diberikan pengguna.
        llm_response_text (str): Respons teks parsial dari LLM.
        rouge_score (float): ROUGE-1 F1 Score.
        total_duration (float): Total durasi proses dalam detik.
        num_epub_pages (int): Jumlah halaman ePub yang diekstrak.
        num_chunks (int): Jumlah chunk teks yang dihasilkan.
        status_message (str): Pesan status akhir proses.
    """
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], PERFORMANCE_LOG_FILE)

    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path) # Muat workbook yang sudah ada
        sheet = workbook.active # Ambil sheet aktif
    else:
        workbook = Workbook() # Buat workbook baru jika belum ada
        sheet = workbook.active
        # Tambahkan header jika workbook baru dibuat
        sheet.append(["Timestamp", "ePub Filename", "LLM Prompt", "LLM Response (Partial)", "ROUGE-1 F1 Score", "Total Duration (s)", "Num ePub Pages", "Num Chunks", "Status Message"])

    # Tambahkan baris data baru
    sheet.append([
        timestamp,
        epub_filename,
        llm_prompt,
        llm_response_text[:100] + "..." if llm_response_text and len(llm_response_text) > 100 else llm_response_text,
        rouge_score,
        total_duration,
        num_epub_pages,
        num_chunks,
        status_message
    ])
    
    try:
        workbook.save(excel_path) # Simpan perubahan ke file Excel
        logging.info(f"Data kinerja dicatat ke: {excel_path}")
    except Exception as e:
        # Tangani error jika file Excel terkunci atau ada masalah izin
        logging.error(f"Gagal menyimpan log kinerja ke '{excel_path}': {e}", exc_info=True)
        logging.error("PENTING: Pastikan file Excel tidak sedang terbuka di program lain!")

def read_performance_log():
    """
    Membaca semua data log kinerja dari file Excel dan mengembalikannya sebagai list of dicts.
    """
    data = []
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], PERFORMANCE_LOG_FILE)
    logging.info(f"Mencoba membaca log kinerja dari: {excel_path}")

    if os.path.exists(excel_path):
        try:
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]] # Ambil header dari baris pertama
            
            # Pastikan ada baris data selain header
            if sheet.max_row > 1:
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)): 
                    # Buat dictionary untuk setiap baris data
                    data.append(dict(zip(headers, row)))
                logging.info(f"Berhasil membaca {len(data)} baris dari log kinerja.")
            else:
                logging.info("File log kinerja ada, tetapi tidak ada data selain header.")
            
        except Exception as e:
            # Tangani error jika file Excel rusak atau terkunci saat dibaca
            logging.error(f"Gagal membaca log kinerja dari '{excel_path}': {e}", exc_info=True)
            logging.warning("Pastikan file Excel tidak rusak atau tidak sedang terbuka.")
            data = [] # Kosongkan data jika ada error saat membaca
    else:
        logging.info("File log kinerja tidak ditemukan, mengembalikan log kosong.")
    return data


# --- Rute Aplikasi Flask ---

@app.route('/')
def index():
    """
    Rute utama untuk menampilkan halaman indeks aplikasi.
    Membaca log kinerja saat halaman dimuat untuk ditampilkan di UI.
    """
    performance_logs = read_performance_log() # Baca log untuk ditampilkan saat halaman dimuat
    return render_template('index.html', performance_logs=performance_logs)

@app.route('/upload', methods=['POST'])
def upload_file():
    """
    Menangani unggahan file ePub, memprosesnya, dan mengembalikan hasil ke frontend.
    Ini adalah alur kerja inti aplikasi.
    """
    start_time = time.time() # Mulai hitung waktu proses end-to-end
    
    # Validasi dasar file yang diunggah
    if 'epub_file' not in request.files:
        logging.error("Tidak ada bagian file dalam permintaan.")
        return jsonify({"error": "Tidak ada file yang diunggah."}), 400
    
    file = request.files['epub_file']
    llm_prompt = request.form.get('llm_prompt', '').strip() 
    
    # Ambil nilai checkbox untuk menentukan apakah halaman ePub asli harus dirender
    render_epub_pages = request.form.get('render_epub_pages') == 'true' 
    
    if file.filename == '':
        logging.warning("Tidak ada file yang dipilih oleh pengguna.")
        return jsonify({"error": "Tidak ada file yang dipilih."}), 400
    
    if file and file.filename.lower().endswith('.epub'):
        original_filename = file.filename
        clean_filename_prefix = image_renderer.clean_filename(os.path.splitext(original_filename)[0])
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], original_filename)
        
        # Buat subfolder unik untuk setiap sesi unggahan di generated_images
        unique_output_subfolder_name = clean_filename_prefix + '_' + str(os.getpid()) + '_' + os.urandom(4).hex() 
        unique_output_full_path = os.path.join(app.config['GENERATED_IMAGES_FOLDER'], unique_output_subfolder_name)
        
        # Direktori sementara untuk ekstraksi aset ePub (untuk base_url Playwright)
        epub_extract_temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'epub_extracts', clean_filename_prefix + '_' + str(os.getpid()))

        # Inisialisasi variabel-variabel untuk hasil dan logging (PENTING: Semua inisialisasi di sini)
        llm_response_text = "N/A" # Default value
        llm_response_image_url = None 
        image_urls = [] # URL gambar halaman ePub asli
        rouge_score = 0.0 # ROUGE score awal
        num_epub_pages_extracted = 0 # Jumlah halaman ePub yang diekstrak
        num_chunks_generated = 0 # Jumlah chunk yang dihasilkan
        llm_prompt_original = llm_prompt # Simpan prompt asli untuk logging
        requested_bg_color_rgb = None # Warna latar belakang yang diminta untuk gambar LLM
        generated_ai_background_path = None # Path gambar latar belakang AI yang digenerate/fallback
        
        status_message = "Processing successful" # Pesan status default untuk log

        # Daftar path ke gambar latar belakang fallback yang sudah didesain
        fallback_bg_dir = os.path.join(app.root_path, 'static', 'images', 'fallback_ai_bgs')
        fallback_bg_images = [os.path.join(fallback_bg_dir, f) for f in os.listdir(fallback_bg_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))]


        try:
            # Simpan file ePub yang diunggah
            file.save(filepath)
            logging.info(f"File '{original_filename}' berhasil diunggah ke '{filepath}'")
            # Buat folder output unik dan folder ekstraksi sementara
            os.makedirs(unique_output_full_path, exist_ok=True) 
            os.makedirs(epub_extract_temp_dir, exist_ok=True) 

            # --- Ekstraksi Konten ePub (HTML, CSS, Gambar Internal) ---
            logging.info(f"Mulai mengekstrak konten dari '{filepath}' ke '{epub_extract_temp_dir}'...")
            html_contents, extracted_asset_paths = epub_processor.extract_epub_content(filepath, epub_extract_temp_dir)
            num_epub_pages_extracted = len(html_contents) # Catat jumlah halaman HTML yang diekstrak
            if not html_contents:
                logging.warning(f"Tidak ada konten HTML yang diekstrak dari '{original_filename}'.")
                raise ValueError("Tidak ada konten yang dapat diekstrak dari ePub ini.")

            # --- Rendering Gambar Konten ePub Asli (Menggunakan Playwright) ---
            if render_epub_pages: 
                logging.info(f"Mulai merender {num_epub_pages_extracted} bagian HTML menjadi gambar menggunakan Playwright...")
                generated_full_paths = image_renderer.render_html_to_images(
                    html_contents, 
                    unique_output_full_path,
                    clean_filename_prefix,
                    base_url=f"file:///{epub_extract_temp_dir.replace(os.sep, '/')}/" # base_url untuk Playwright
                )
                
                # Konversi path gambar lokal menjadi URL yang bisa diakses web
                for full_path in generated_full_paths:
                    image_filename_only = os.path.basename(full_path)
                    subfolder_and_filename = os.path.join(os.path.basename(unique_output_full_path), image_filename_only).replace("\\", "/")
                    image_urls.append(f"/generated_images/{subfolder_and_filename}")
            else:
                logging.info("Rendering gambar halaman ePub asli dilewati sesuai permintaan pengguna.")
                image_urls = [] # Pastikan list URL gambar kosong jika rendering dilewati

            # --- Pemrosesan LLM dan Generasi Gambar AI ---
            if llm_prompt: 
                # Ekstrak permintaan warna latar belakang dari prompt asli pengguna
                llm_prompt_cleaned_for_llm, requested_bg_color_rgb = extract_background_color_from_prompt(llm_prompt_original)

                logging.info(f"Mulai memproses prompt LLM: '{llm_prompt_cleaned_for_llm}'")
                
                full_epub_text = " ".join([BeautifulSoup(html, 'html.parser').get_text(separator=' ', strip=True) for html in html_contents])
                
                # --- Implementasi Chunking ---
                chunks = epub_processor.split_text_into_chunks(full_epub_text, max_len=1500) 
                num_chunks_generated = len(chunks) 
                logging.info(f"Teks ePub dipecah menjadi {num_chunks_generated} chunk.")

                # --- Pemilihan Konteks untuk LLM (RAG Dasar) ---
                num_chunks_for_context = min(5, len(chunks)) 
                context_for_llm = "\n\n".join(chunks[:num_chunks_for_context])
                logging.info(f"Menggunakan {num_chunks_for_context} chunk sebagai konteks untuk LLM.")
                
                final_llm_prompt = f"Teks dari buku ePub (bagian awal) adalah:\n\n---\n{context_for_llm}\n---\n\nBerdasarkan teks di atas, {llm_prompt_cleaned_for_llm}\n\nJANGAN sertakan format HTML, Markdown, atau styling apapun dalam respons Anda. Hanya berikan teks murni."
                
                llm_response_text = llm_integrator.get_gemini_response(final_llm_prompt) 
                logging.info(f"Respons LLM diterima: {llm_response_text[:100]}...")

                # --- Hitung ROUGE Score ---
                rouge_score = hitung_rouge_score(llm_prompt_cleaned_for_llm, llm_response_text)
                logging.info(f"ROUGE-1 F1 Score (Prompt vs Response): {rouge_score}")


                # --- GENERASI GAMBAR AI (LATAR BELAKANG) atau FALLBACK ---
                if not requested_bg_color_rgb: # Hanya coba generate AI jika tidak ada warna spesifik yang diminta
                    image_gen_prompt = f"Minimalist abstract background, simple elegant shapes, soft warm colors, digital art. Related to the theme of: '{llm_response_text[:200].replace('\n', ' ')}' --v 5.2 --style raw" 
                    
                    ai_image_output_filename = f"{clean_filename_prefix}_ai_bg.png"
                    ai_image_full_path = os.path.join(unique_output_full_path, ai_image_output_filename)

                    logging.info(f"Mulai generasi gambar AI untuk latar belakang: '{image_gen_prompt[:100]}...'")
                    generated_ai_background_path = llm_integrator.generate_image_from_text(
                        image_gen_prompt, ai_image_full_path
                    )
                    
                    if not generated_ai_background_path:
                        logging.warning("Gagal generate gambar AI. Mencoba menggunakan gambar latar belakang fallback yang sudah didesain.")
                        if fallback_bg_images: # <--- Penggunaan fallback_bg_images
                            generated_ai_background_path = random.choice(fallback_bg_images)
                            logging.info(f"Menggunakan gambar fallback: {os.path.basename(generated_ai_background_path)}")
                        else:
                            logging.warning("Tidak ada gambar latar belakang fallback yang ditemukan. Membuat gambar default polos.")
                            temp_default_bg_path = os.path.join(unique_output_full_path, "default_plain_bg.png")
                            Image.new('RGB', (800, 400), (240, 240, 240)).save(temp_default_bg_path) 
                            generated_ai_background_path = temp_default_bg_path
                else:
                    logging.info(f"Warna latar belakang spesifik diminta ({requested_bg_color_rgb}). Melewatkan generasi gambar AI.")


                # --- Render Respons LLM ke Gambar yang Didesain dengan Pillow ---
                if llm_response_text and llm_response_text != "Tidak ada respons yang dihasilkan dari model.":
                    llm_image_filename = f"{clean_filename_prefix}_llm_result.png"
                    llm_image_full_path = os.path.join(unique_output_full_path, llm_image_filename)
                    
                    logging.info(f"Merender respons LLM ke gambar yang didesain: '{llm_image_filename}'")
                    
                    font_for_pillow_render = os.path.join(app.root_path, 'fonts', 'NotoSansArabic-Regular.ttf')

                    rendered_llm_image_path = image_renderer.render_llm_text_to_designed_image(
                        llm_response_text, 
                        llm_image_full_path, 
                        font_path=font_for_pillow_render,
                        ai_background_path=generated_ai_background_path, 
                        requested_bg_color=requested_bg_color_rgb 
                    )

                    if rendered_llm_image_path:
                        llm_response_image_url = f"/generated_images/{os.path.basename(unique_output_full_path)}/{os.path.basename(rendered_llm_image_path)}".replace("\\", "/")
                        logging.info(f"Respons LLM berhasil dirender ke gambar: {llm_response_image_url}")
                    else:
                        logging.error("Gagal merender gambar hasil LLM dengan Pillow.")
                        llm_response_image_url = None
                else:
                    logging.warning("Respons LLM kosong atau tidak valid, tidak merender gambar hasil LLM.")
                    llm_response_text = "Tidak ada respons dari AI." 

            # --- Pembersihan File Sementara ---
            os.remove(filepath) 
            logging.info(f"File ePub '{original_filename}' dihapus dari folder unggahan.")
            
            if os.path.exists(epub_extract_temp_dir):
                shutil.rmtree(epub_extract_temp_dir) 
                logging.info(f"Folder ekstraksi sementara '{epub_extract_temp_dir}' dihapus.")

            # Perbarui pesan sukses yang akan ditampilkan di frontend
            final_message = f"Berhasil mengkonversi '{original_filename}'. "
            if render_epub_pages and len(image_urls) > 0: 
                final_message += f"Dihasilkan {len(image_urls)} gambar konten ePub."
            elif render_epub_pages and len(image_urls) == 0: 
                final_message += "Tidak ada gambar konten ePub yang dihasilkan (cek log server)."
            else: 
                final_message += "Rendering gambar konten ePub dilewati."

            # Catat data kinerja ke file Excel
            end_time = time.time()
            total_duration = round(end_time - start_time, 2)
            log_performance_data(
                time.strftime("%Y-%m-%d %H:%M:%S"),
                original_filename,
                llm_prompt_original, 
                llm_response_text,
                rouge_score,
                total_duration,
                num_epub_pages_extracted,
                num_chunks_generated,
                status_message
            )

            # Baca ulang log kinerja setelah dicatat untuk dikirim ke frontend
            updated_performance_logs = read_performance_log()


            return jsonify({
                "message": final_message, 
                "image_urls": image_urls,
                "llm_response_text": llm_response_text,
                "llm_image_url": llm_response_image_url, 
                "performance_log": updated_performance_logs 
            }), 200

        except Exception as e:
            logging.error(f"Error saat memproses file '{original_filename}': {e}", exc_info=True)
            status_message = f"Failed: {str(e)}"
            
            end_time = time.time()
            total_duration = round(end_time - start_time, 2)
            # Pastikan semua variabel yang digunakan dalam log_performance_data terdefinisi
            log_performance_data(
                time.strftime("%Y-%m-%d %H:%M:%S"),
                original_filename,
                llm_prompt_original if 'llm_prompt_original' in locals() else "N/A", # Menggunakan prompt asli jika tersedia
                llm_response_text if 'llm_response_text' in locals() else "Error during processing", # Menggunakan respons LLM jika tersedia
                rouge_score if 'rouge_score' in locals() else 0.0, # Menggunakan ROUGE score jika tersedia
                total_duration,
                num_epub_pages_extracted if 'num_epub_pages_extracted' in locals() else 0, 
                num_chunks_generated if 'num_chunks_generated' in locals() else 0, 
                status_message
            )
            # Baca ulang log kinerja setelah dicatat error untuk dikirim ke frontend
            updated_performance_logs = read_performance_log()

            # Pembersihan file dan folder yang mungkin tersisa jika terjadi error
            if os.path.exists(filepath):
                os.remove(filepath)
            if os.path.exists(unique_output_full_path):
                shutil.rmtree(unique_output_full_path)
            if os.path.exists(epub_extract_temp_dir):
                shutil.rmtree(epub_extract_temp_dir)
            
            # Mengembalikan respons error JSON ke frontend
            return jsonify({
                "error": f"Gagal memproses file: {str(e)}. Cek log server untuk detail.",
                "performance_log": updated_performance_logs 
            }), 500
    else:
        # Menangani unggahan file dengan format yang tidak didukung
        logging.warning(f"File '{file.filename}' yang diunggah bukan format .epub atau tidak valid.")
        return jsonify({"error": "Format file tidak didukung. Harap unggah file .epub."}), 400

# Rute untuk melayani gambar yang dihasilkan dari subfolder unik
@app.route('/generated_images/<subfolder>/<filename>')
def serve_generated_image(subfolder, filename):
    """
    Melayani file gambar yang dihasilkan dari subfolder unik.
    Ini memungkinkan gambar diakses melalui URL di browser.
    """
    full_path_to_subfolder = os.path.join(app.config['GENERATED_IMAGES_FOLDER'], subfolder)
    # send_from_directory secara aman melayani file dari direktori yang ditentukan
    return send_from_directory(full_path_to_subfolder, filename)

# Rute untuk mengunduh log kinerja
@app.route('/download-performance-log')
def download_performance_log():
    """
    Memungkinkan pengguna untuk mengunduh file log kinerja dalam format Excel.
    """
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], PERFORMANCE_LOG_FILE)
    if os.path.exists(excel_path):
        return send_from_directory(app.config['UPLOAD_FOLDER'], PERFORMANCE_LOG_FILE, as_attachment=True)
    else:
        return "Log kinerja belum tersedia.", 404

# Rute untuk menghapus log kinerja
@app.route('/clear-performance-log', methods=['POST'])
def clear_performance_log():
    """
    Menghapus file log kinerja Excel dari server.
    """
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], PERFORMANCE_LOG_FILE)
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)
            logging.info(f"Log kinerja '{excel_path}' berhasil dihapus.")
            # Setelah menghapus, kirim log kosong ke frontend
            return jsonify({"message": "Log kinerja berhasil dihapus.", "performance_log": []}), 200
        except Exception as e:
            # Tangani error jika file terkunci atau ada masalah izin saat menghapus
            logging.error(f"Gagal menghapus log kinerja '{excel_path}': {e}", exc_info=True)
            # Jika gagal, baca ulang log yang ada (mungkin terkunci) dan kirim
            return jsonify({"error": f"Gagal menghapus log: {str(e)}", "performance_log": read_performance_log()}), 500
    else:
        logging.info("Tidak ada log kinerja untuk dihapus.")
        return jsonify({"message": "Tidak ada log kinerja untuk dihapus.", "performance_log": []}), 200


# Menjalankan Aplikasi Flask
if __name__ == '__main__':
    # app.run(debug=True) akan menjalankan server pengembangan Flask
    # debug=True akan memberikan pesan error yang lebih detail dan reload otomatis saat kode berubah
    app.run(debug=True, port=5000)
